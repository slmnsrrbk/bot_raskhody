import datetime
import io
import os
import sqlite3
import tempfile
import unittest

os.environ.setdefault("TELEGRAM_TOKEN", "test")
import ai  # noqa: E402
import crypto  # noqa: E402
import export  # noqa: E402
import storage  # noqa: E402
from openpyxl import load_workbook  # noqa: E402


class CryptoTests(unittest.TestCase):
    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        crypto.UNLOCK_DIR = storage.Path(self.tmp.name) / "shm"
        crypto.reset_cache()
        storage.DB_FILE = storage.Path(self.tmp.name) / "t.db"
        storage.init_db()
        for uid in (1, 2, 5, 7):
            crypto.setup_pin(uid, f"pass{uid}00")

    def tearDown(self):
        self.tmp.cleanup()
        crypto.reset_cache()

    def test_roundtrip_and_per_user_keys(self):
        c = crypto.encrypt(1, "Кофе")
        self.assertTrue(c.startswith("enc2:"))
        self.assertEqual(crypto.decrypt(1, c), "Кофе")
        with self.assertRaises(Exception):
            crypto.decrypt(2, c)  # чужим ключом не расшифровать
        self.assertNotEqual(crypto.encrypt(1, "Кофе"), c)  # случайный nonce

    def test_lock_unlock_and_wrong_pin(self):
        storage.add_expense(1, "Кофе", 250, "Еда", datetime.date(2026, 9, 3))
        crypto.lock(1)
        self.assertEqual(crypto.status(1), "locked")
        with self.assertRaises(crypto.Locked):
            storage.list_expenses(1)
        with self.assertRaises(crypto.WrongPin):
            crypto.unlock(1, "wrong-pass")
        crypto.unlock(1, "pass100")
        self.assertEqual(storage.list_expenses(1)[0]["name"], "Кофе")
        # ключ доступен другому «процессу» через tmpfs-файл
        crypto.reset_cache()
        self.assertEqual(crypto.status(1), "unlocked")
        self.assertEqual(oct(crypto._key_path(1).stat().st_mode & 0o777), "0o600")

    def test_no_secrets_on_disk_without_pin(self):
        storage.add_expense(1, "Секрет", 999, "Еда", datetime.date(2026, 9, 3))
        crypto.lock(1)
        crypto.reset_cache()
        raw = sqlite3.connect(storage.DB_FILE)
        blob = " ".join(str(v) for row in raw.execute("SELECT * FROM expenses") for v in row)
        blob += " ".join(str(v) for row in raw.execute("SELECT * FROM user_keys") for v in row)
        self.assertNotIn("Секрет", blob)
        self.assertNotIn("999", blob)
        self.assertNotIn("pass100", blob)
        self.assertFalse(list(crypto.UNLOCK_DIR.glob("1.key")))

    def test_change_pin_and_attempts(self):
        storage.add_expense(1, "Кофе", 250, "Еда", datetime.date(2026, 9, 3))
        crypto.change_pin(1, "pass100", "newpass1")
        crypto.lock(1)
        crypto.unlock(1, "newpass1")
        self.assertEqual(storage.list_expenses(1)[0]["amount"], 250)
        crypto.lock(1)
        for _ in range(crypto.MAX_ATTEMPTS):
            with self.assertRaises(crypto.WrongPin):
                crypto.unlock(1, "bad-bad-bad")
        with self.assertRaises(crypto.TooManyAttempts):
            crypto.unlock(1, "newpass1")
        with self.assertRaises(ValueError):
            crypto.setup_pin(9, "123")  # слишком короткий

    def test_wipe_user(self):
        storage.add_expense(1, "Кофе", 250, "Еда", datetime.date(2026, 9, 3))
        storage.wipe_user(1)
        self.assertEqual(crypto.status(1), "nopin")
        crypto.setup_pin(1, "another1")
        self.assertEqual(storage.list_expenses(1), [])

    def test_db_has_no_plaintext(self):
        storage.add_expense(5, "Секретная покупка", 1234, "Еда", datetime.date(2026, 9, 3))
        storage.set_limits(5, daily=2500)
        raw = sqlite3.connect(storage.DB_FILE)
        blob = " ".join(str(v) for row in raw.execute("SELECT name, amount, category FROM expenses") for v in row)
        self.assertNotIn("Секретная", blob)
        self.assertNotIn("1234", blob)
        self.assertNotIn("Еда", blob)
        lim = raw.execute("SELECT daily FROM limits").fetchone()[0]
        self.assertTrue(lim.startswith("enc2:"))
        self.assertEqual(storage.list_expenses(5)[0]["name"], "Секретная покупка")
        self.assertEqual(storage.get_limits(5)["daily"], 2500)

    def test_legacy_rows_are_removed(self):
        raw = sqlite3.connect(storage.DB_FILE)
        raw.execute("INSERT INTO expenses(user_id,name,amount,category,date,created_at) VALUES(7,'enc1:xxx','enc1:y','enc1:z','2026-09-01','x')")
        raw.commit(); raw.close()
        storage.init_db()
        self.assertEqual(storage.list_expenses(7), [])

    def test_wipe_and_bulk(self):
        added = storage.add_expenses_bulk(1, [{"name": "a", "amount": 1, "category": "Еда", "date": "2026-09-01"},
                                             {"name": "b", "amount": 2, "category": "Другое", "date": "2026-09-01"}])
        self.assertEqual(len(added), 2)
        self.assertEqual(storage.delete_many(1, [added[0]["id"], 999]), 1)
        storage.wipe_all()
        self.assertEqual(storage.list_expenses(1), [])


class ExportTests(unittest.TestCase):
    def test_xlsx(self):
        items = [{"id": 1, "name": "Кофе", "amount": 250, "category": "Еда", "iso": "2026-09-03", "date": "03.09.2026"},
                 {"id": 2, "name": "Такси", "amount": 480, "category": "Транспорт", "iso": "2026-08-30", "date": "30.08.2026"}]
        data = export.build_xlsx(items, "Месяц", datetime.date(2026, 9, 3))
        wb = load_workbook(io.BytesIO(data))
        self.assertEqual(wb.sheetnames, ["Расходы", "По категориям", "По месяцам", "Инфо"])
        ws = wb["Расходы"]
        self.assertEqual([c.value for c in ws[1]], ["Дата", "Название", "Категория", "Сумма"])
        self.assertEqual(ws["B2"].value, "Такси")  # по возрастанию даты
        self.assertEqual(ws["D4"].value, None)
        self.assertEqual(ws["D5"].value, 730)
        self.assertEqual(wb["По категориям"]["A2"].value, "Транспорт")
        self.assertEqual(wb["По месяцам"].max_row, 3)
        self.assertEqual(export.filename("30", datetime.date(2026, 9, 3)), "расходы_месяц_2026-09-03.xlsx")


class ReceiptCleanTests(unittest.TestCase):
    def test_clean(self):
        parsed = ai._clean_receipt({"store": "Пятёрочка", "date": "03.09.2026", "total": "517.77", "items": [
            {"name": "Молоко", "amount": "79,90", "category": "Еда"},
            {"name": "Пакет", "amount": 7, "category": "нет такой"},
            {"name": "", "amount": 5}, {"name": "x", "amount": "abc"}, "мусор"]})
        self.assertEqual(parsed["store"], "Пятёрочка")
        self.assertEqual(parsed["date"], "2026-09-03")
        self.assertEqual(parsed["total"], 518)
        self.assertEqual([(i["name"], i["amount"], i["category"]) for i in parsed["items"]], [("Молоко", 80, "Еда"), ("Пакет", 7, "Другое")])

    def test_extract_json(self):
        self.assertEqual(ai._extract_json('```json\n{"items": []}\n```'), {"items": []})
        self.assertEqual(ai._extract_json('Вот ответ: {"items": [{"name":"a","amount":1}]} готово')["items"][0]["name"], "a")
        self.assertIsNone(ai._extract_json("не json"))


if __name__ == "__main__":
    unittest.main()
