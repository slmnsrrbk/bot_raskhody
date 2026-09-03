"""Выгрузка расходов в Excel (openpyxl)."""
import datetime
import io
from collections import OrderedDict

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

MONTHS = ["январь", "февраль", "март", "апрель", "май", "июнь", "июль", "август", "сентябрь", "октябрь", "ноябрь", "декабрь"]
PERIODS = OrderedDict([("7", ("Неделя", 7)), ("30", ("Месяц", 30)), ("90", ("3 месяца", 90)), ("365", ("Год", 365)), ("all", ("Всё время", None))])
MONEY = '#,##0 "₽"'


def money_format(symbol: str = "₽") -> str:
    return '#,##0 "' + str(symbol).replace('"', "") + '"'
HEAD_FILL = PatternFill("solid", fgColor="F2F2F7")


def _header(ws, titles):
    ws.append(titles)
    for c in ws[1]:
        c.font = Font(bold=True)
        c.fill = HEAD_FILL
        c.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"


def _autosize(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_xlsx(items, period_label: str, today: datetime.date, symbol: str = "₽") -> bytes:
    """items — список записей storage.list_expenses (новые первыми); symbol — валюта пользователя."""
    MONEY = money_format(symbol)  # noqa: N806 — локальный формат для этой книги
    wb = Workbook()
    ws = wb.active
    ws.title = "Расходы"
    _header(ws, ["Дата", "Название", "Категория", "Сумма", "Заметка", "Исходная сумма"])
    total = 0
    for it in sorted(items, key=lambda x: (x["iso"], x["id"])):
        d = datetime.date.fromisoformat(it["iso"])
        orig = f"{it['orig_amount']} {it['currency']}" if it.get("currency") and it.get("orig_amount") != it["amount"] else ""
        ws.append([d, it["name"], it["category"], it["amount"], it.get("note") or "", orig])
        ws.cell(ws.max_row, 1).number_format = "DD.MM.YYYY"
        ws.cell(ws.max_row, 4).number_format = MONEY
        total += it["amount"]
    ws.append([])
    ws.append(["Итого", "", "", total])
    ws.cell(ws.max_row, 1).font = Font(bold=True)
    ws.cell(ws.max_row, 4).font = Font(bold=True)
    ws.cell(ws.max_row, 4).number_format = MONEY
    if items:
        ws.auto_filter.ref = f"A1:D{len(items) + 1}"
    _autosize(ws, [12, 36, 16, 14, 30, 16])

    cats = {}
    for it in items:
        cats[it["category"]] = cats.get(it["category"], 0) + it["amount"]
    ws2 = wb.create_sheet("По категориям")
    _header(ws2, ["Категория", "Сумма", "Доля"])
    for cat, amt in sorted(cats.items(), key=lambda kv: -kv[1]):
        ws2.append([cat, amt, amt / total if total else 0])
        ws2.cell(ws2.max_row, 2).number_format = MONEY
        ws2.cell(ws2.max_row, 3).number_format = "0.0%"
    ws2.append(["Итого", total, 1 if total else 0])
    ws2.cell(ws2.max_row, 1).font = Font(bold=True)
    ws2.cell(ws2.max_row, 2).number_format = MONEY
    ws2.cell(ws2.max_row, 3).number_format = "0.0%"
    _autosize(ws2, [18, 14, 10])

    months = {}
    for it in items:
        months[it["iso"][:7]] = months.get(it["iso"][:7], 0) + it["amount"]
    ws3 = wb.create_sheet("По месяцам")
    _header(ws3, ["Месяц", "Сумма"])
    for ym in sorted(months):
        y, m = ym.split("-")
        ws3.append([f"{MONTHS[int(m) - 1].capitalize()} {y}", months[ym]])
        ws3.cell(ws3.max_row, 2).number_format = MONEY
    _autosize(ws3, [18, 14])

    ws4 = wb.create_sheet("Инфо")
    ws4.append(["Период", period_label])
    ws4.append(["Сформировано", today.strftime("%d.%m.%Y")])
    ws4.append(["Записей", len(items)])
    ws4.append(["Итого", total])
    ws4.cell(4, 2).number_format = MONEY
    _autosize(ws4, [16, 24])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def filename(period_key: str, today: datetime.date) -> str:
    label = PERIODS.get(period_key, ("Период", None))[0].lower().replace(" ", "_")
    return f"расходы_{label}_{today.strftime('%Y-%m-%d')}.xlsx"
